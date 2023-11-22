from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font, Color
from config import get_config
import sys
from tqdm import tqdm
from time import time


def autoresize_columns(worksheet: Workbook.worksheets, starting_column=None, ending_column=None, column_width=20):
    """ The width of each column is adjusted.
    The function consider the largest width of the cell within a columns
    The starting column cannot be 0
    """
    last_col: int = worksheet.max_column + 1
    first_col: int = worksheet.min_column
    if starting_column is not None:
        if starting_column == 0:
            raise Exception('Starting column cannot start with 0')
        first_col = starting_column
    if ending_column is not None:
        if ending_column > last_col:
            raise Exception(f'Ending column cannot be bigger than {last_col}')
        last_col = ending_column
    total = last_col-first_col
    bar: tqdm = tqdm(total=total, disable=False)
    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(first_col, last_col):
        bar.update()
        col_header = list(worksheet.columns)[col-1][0]
        header_text = str(col_header.value)
        if get_config()['output_hours_num_col'].strip('()') in header_text or get_config()['output_student_num_col'].strip('()') in header_text:
            width = max(1, int(column_width/4))
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col, width=width)
            col_header.value = int(format_statistics_column_header(header_text))
            continue
        dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col, width=column_width)
    bar.close()
    worksheet.column_dimensions = dim_holder
    print('='*5+'DIMENSIONS OF COLUMNS HAVE BEEN RESIZED'+'='*5)


def adjust_text_alignment(worksheet: Workbook.worksheets):
    font_size = get_config()['font_size']
    header_font_color = Color(rgb=formatted_color('#ffffffff'))
    row_count: int = worksheet.max_row
    column_count: int = worksheet.max_column
    total: int = row_count*column_count
    side = Side(border_style=None)
    no_border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side,
                         right=thin_side,
                         top=thin_side,
                         bottom=thin_side)
    start_time = time()
    bar: tqdm = tqdm(total=total, disable=False)
    col_num = 0
    for col in worksheet.columns:
        col_num += 1
        row_num = 0
        for cell in col:
            row_num += 1
            bar.update()
            """=== Cell Conditions ==="""

            is_public_holiday: bool = (str(cell.value) == 'PH')
            is_leave: bool = (str(cell.value) == 'L')
            is_booking: bool = (not is_public_holiday
                                and not is_leave
                                and cell.value is not None
                                and col_num > get_config()['freeze_columns'])
            is_column_header: bool = (row_num == 1)

            """ === Set text alignment === """
            if is_column_header:
                worksheet[cell.coordinate].alignment = Alignment(text_rotation=90,
                                                                 vertical='center',
                                                                 horizontal='center')
                set_column_header_color(worksheet, cell)
                set_font(cell, size=font_size, color=header_font_color)
                continue
            worksheet[cell.coordinate].alignment = Alignment(wrap_text=True, horizontal='center')

            """ === Set border ==="""
            cell.border = no_border

            """ === Set font style ==="""
            set_font(cell, size=font_size)

            """=== Set cell color ==="""
            if row_num % 2 == 0:
                set_default_color(worksheet, cell)

            """=== Set cell color by value ==="""
            if is_public_holiday:
                set_public_holiday_color(worksheet, cell)
            elif is_leave:
                set_leave_color(worksheet, cell)
            elif is_booking:
                set_booking_color(worksheet, cell)
                cell.border = thin_border
    bar.close()
    print('='*5+f'CELL STYLE HAS BEEN ADDED'+5*'=')
    print('Time: ', time()-start_time)


def formatted_color(color: str):
    return color.replace('#', '').upper()


def format_statistics_column_header(text: str) -> str:
    # keep what is inside the round brackets ()
    text = text.strip('()')
    # keep only numbers after semicolon :
    text = text.split(':')[1]
    return text.strip()


def update_cell_text(cell, new_value: str) -> None:
    cell.value = new_value


def convert_to_column_name(column_number: int) -> str:
    column_name = ""
    while column_number > 0:
        remainder = (column_number - 1) % 26
        column_name = chr(65 + remainder) + column_name
        column_number = (column_number - 1) // 26
    return column_name


def set_public_holiday_color(worksheet, cell) -> None:
    color: str = get_config()['cell_colors']['public_holiday']
    color = formatted_color(color)
    fill = PatternFill(start_color=color,
                       end_color=color,
                       fill_type='solid')
    worksheet[cell.coordinate].fill = fill


def set_booking_color(worksheet, cell, color=get_config()['cell_colors']['booking']) -> None:
    color = formatted_color(color)
    fill = PatternFill(start_color=color,
                       end_color=color,
                       fill_type='solid')
    worksheet[cell.coordinate].fill = fill


def set_leave_color(worksheet, cell) -> None:
    color: str = get_config()['cell_colors']['leave']
    color = formatted_color(color)
    fill = PatternFill(start_color=color,
                       end_color=color,
                       fill_type='solid')
    worksheet[cell.coordinate].fill = fill


def set_default_color(worksheet, cell) -> None:
    color: str = get_config()['cell_colors']['default']
    color = formatted_color(color)
    fill = PatternFill(start_color=color,
                       end_color=color,
                       fill_type='solid')
    worksheet[cell.coordinate].fill = fill


def set_column_header_color(worksheet, cell) -> None:
    color: str = get_config()['cell_colors']['header']
    color = formatted_color(color)
    fill = PatternFill(start_color=color,
                       end_color=color,
                       fill_type='solid')
    worksheet[cell.coordinate].fill = fill


def freeze(worksheet, columns: int, rows: int) -> None:
    if rows <= 0 or columns <= 0:
        print('Incorrect number of rows or columns entered', file=sys.stderr)
        return None
    columns += 1
    rows += 1
    column = convert_to_column_name(columns)
    worksheet.freeze_panes = column+str(rows)
    print(f'{columns} columns and {rows} rows have been freeze')


def set_font(cell, size: int, is_bold=False, color=None) -> None:
    font = Font(size=str(size), bold=is_bold, color=color)
    cell.font = font
